import pandas as pd
from openpyxl import Workbook
from utils.recover_files import convert_with_excel
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import time
import re
from datetime import datetime
from utils.convert_style import redactor, redactor_ws

start_time = time.perf_counter()
current_date = datetime.now().strftime('%d-%m-%Y')



file = 'test.xlsx'
file_2 = 'sales.xlsx'

try:
    df = pd.read_excel(file, skiprows=10, engine='openpyxl')
except Exception as e:
    convert_with_excel(file)
    df = pd.read_excel(file, skiprows=10, engine='openpyxl')
dlina = len(df)
print(f'открыт файл с остатками на {dlina} позиций')
names = ['Завенягина', 'Маркса', 'ТК ДжазМолл', 'Ленина', 'Склад']
df.columns = ['', 'Номенклатура'] + names
df = df.drop('', axis=1)
df.iloc[:, 1:] = df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce').fillna(0) # заменяем Nan на 0

def clean_nomenclature(x):
    if isinstance(x, str):
        return re.sub(r', [^,]*$', '', x)
    return x


try:
    df2 = pd.read_excel(file_2, skiprows=8, engine='openpyxl')
    print(f'открыт файл с продажами')
    df2.columns = ['', 'Номенклатура', 'Продажи']
    #df2['Номенклатура'] = df2['Номенклатура'].apply(clean_nomenclature)
    df2 = df2.drop('', axis=1)
except FileNotFoundError:
    df2 = pd.DataFrame(columns=['Номенклатура', 'Продажи'])
    print(f'отсутствует файл с продажами')
except Exception as e:
    convert_with_excel(file_2)
    df2 = pd.read_excel(file_2, skiprows=10, engine='openpyxl')
    print(f'открыт файл с продажами')
    df2.columns = ['', 'Номенклатура', 'Продажи']
    #df2['Номенклатура'] = df2['Номенклатура'].apply(clean_nomenclature)
    df2 = df2.drop('', axis=1)



df = pd.merge(df, df2, on='Номенклатура', how='left')
df['Продажи'] = pd.to_numeric(df['Продажи'], errors='coerce').fillna(0)

prioritet = ['Ленина', 'Завенягина', 'ТК ДжазМолл']
df = df.reindex(columns=['Номенклатура', 'Продажи', "Маркса", 'Склад'] + prioritet)

conditions = [
    df['Продажи'] - df['Маркса'] > 0,
    df['Продажи'] - df['Маркса'] < 0,
    df['Продажи'] - df['Маркса'] == 0
]

choices = [
    df['Продажи'] - df['Маркса'],  # если положительное
    0,  # если отрицательное
    1  # если ноль
]

df.insert(2, 'Рекомендовано к заказу', np.select(conditions, choices, 0))
df['ordered'] = False

df = df[(df['Рекомендовано к заказу'] > 0)]

print(f'Сформирована рекомендация к заказу')

print(f'обработка остатков на складе')

for idx in df.index:
    if df.loc[idx, "Склад"] >= df.loc[idx, "Рекомендовано к заказу"]:
        df.loc[idx, 'ordered'] = True
    elif 0 < df.loc[idx, "Склад"] < df.loc[idx, "Рекомендовано к заказу"]:
        df.loc[idx, "Рекомендовано к заказу"] = df.loc[idx, "Склад"]
        df.loc[idx, 'ordered'] = True

result = df[(df['ordered'] == True)]
if not result.empty:
    result[['Номенклатура', "Рекомендовано к заказу", 'Маркса', 'Склад']].to_excel(f'заказы со склада от {current_date}.xlsx', index=False)
    redactor(f'заказы со склада от {current_date}.xlsx')
    print(f"создан файл 'заказы со склада от {current_date}.xlsx' найдено {len(result)} позиций")
else:
    print("Нет доступных позиций к заказу со склада")

df = df[(df['ordered'] == False)]
df.insert(2, 'Заказ из магазина', 0)
df['ordered'] = 0


for idx in df.index:  # Перебираем строки DataFrame
    # Ищем первый доступный склад в порядке приоритета

    for i in range(len(prioritet)):  # i - индекс в списке складов
        warehouse = prioritet[i]  # Получаем название склада

        if df.loc[idx, warehouse] > 1 and df.loc[idx, "Рекомендовано к заказу"] > df.loc[idx, "ordered"]:  # Проверяем наличие товара
            # Устанавливаем 1 на найденном складе
            df.loc[idx, "ordered"] += 1
            df.loc[idx, "Заказ из магазина"] = 1
            # Устанавливаем 0 на всех остальных складах (после найденного)
            # for j in range(i + 1, len(prioritet)):
            #     df.loc[idx, prioritet[j]] = 0

        else:
            df.loc[idx, warehouse] = 0

print(f'обработка завершена, приступаем к созданию отчетов')

# Создаем новый workbook
wb = Workbook()
wb.remove(wb.active)

# Создаем листы для каждого склада
for sklad in prioritet:
    # Фильтруем строки, где этот склад имеет значение 1
    otchet = df[(df[sklad] > 1)]

    # Проверяем, есть ли данные для этого склада
    if not otchet.empty:
        # Создаем новый лист с названием склада
        ws = wb.create_sheet(title=sklad)

        # Добавляем данные
        for r in dataframe_to_rows(otchet[['Номенклатура',"Заказ из магазина", "Рекомендовано к заказу", "Маркса",  sklad, "ordered"]], index=False,
                                   header=True):
            ws.append(r)

        redactor_ws(ws)
        print(f"Создан лист '{sklad}' найдено {len(otchet)} позиций")
    else:
        print(f"Нет данных для склада {sklad}")

# Сохраняем файл
filename = f'заказы с магазинов от {current_date}.xlsx'
wb.save(filename)
print(f"Создан файл '{filename}' с {len(wb.sheetnames)} листами")

df = df[(df['ordered'] < df["Рекомендовано к заказу"])]
df = df[(df['Продажи'] > 0)]
df[['Номенклатура', 'Продажи', 'ordered']].to_excel(f'дефицит от {current_date}.xlsx', index=False)
redactor(f'дефицит от {current_date}.xlsx')
print(f'Cоздан файл - дефицит от {current_date}.xlsx')

end_time = time.perf_counter()
execution_time = end_time - start_time
print(f'Отчет завершен, обработано {dlina} позиций за {execution_time:.4f} секунд')