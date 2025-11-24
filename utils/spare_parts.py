import asyncio
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import time
import re
from datetime import datetime
from utils.convert_style import redactor, redactor_ws
from utils.recover_files import convert_with_excel


async def zip_otchet():
    global df
    start_time = time.perf_counter()
    current_date = datetime.now().strftime('%d-%m-%Y')
    output_dir = Path('C:/MyProjects/1c_scripts/отчеты')
    output_dir.mkdir(exist_ok=True)
    file = Path('C:/MyProjects/1c_scripts/остатки') / 'запчасти.xlsx'
    file_2 = Path('C:/MyProjects/1c_scripts/остатки') / 'sales.xlsx'
    try:
        df = pd.read_excel(file, skiprows=10, engine='openpyxl')
    except Exception as e:
        convert_with_excel(file, file)
        df = pd.read_excel(file, skiprows=10, engine='openpyxl')
    dlina = len(df)
    print(f'открыт файл с отстатками на {dlina} позиций')
    names = ['Артиллерийская', 'Златоуст', 'Златоуст ТРК Тарелка', 'Копейск', 'Завенягина', 'Маркса', 'ТК ДжазМолл',
             'Миасс', 'Гагарина', 'Комсомольский', 'Молодогвардейцев', 'КС Теплотех', 'Ленина',
             'Сталеваров', 'Худякова', 'Склад']
    df.columns = ['', 'Номенклатура'] + names
    df = df.drop('', axis=1)
    df.iloc[:, 1:] = df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce').fillna(0)  # заменяем Nan на 0

    def clean_nomenclature(x):
        if isinstance(x, str):
            return re.sub(r', [^,]*$', '', x).strip()
        return x

    # df['Номенклатура'] = df['Номенклатура'].apply(clean_nomenclature)

    try:
        df2 = pd.read_excel(file_2, skiprows=8, engine='openpyxl')
        print(f'открыт файл с продажами')
        df2.columns = ['', 'Номенклатура', 'Продажи']
        # df2['Номенклатура'] = df2['Номенклатура'].apply(clean_nomenclature)
        df2 = df2.drop('', axis=1)
    except FileNotFoundError:
        df2 = pd.DataFrame(columns=['Номенклатура', 'Продажи'])
        print(f'отсутствует файл с продажами')
    except Exception as e:
        convert_with_excel(file_2, file_2)
        df2 = pd.read_excel(file_2, skiprows=10, engine='openpyxl')
        print(f'открыт файл с продажами')
        df2.columns = ['', 'Номенклатура', 'Продажи']
        # df2['Номенклатура'] = df2['Номенклатура'].apply(clean_nomenclature)
        df2 = df2.drop('', axis=1)

    df = pd.merge(df, df2, on='Номенклатура', how='left')
    df['Продажи'] = pd.to_numeric(df['Продажи'], errors='coerce').fillna(0)

    prioritet = ['Завенягина', 'ТК ДжазМолл', 'Миасс Макеева', 'Миасс', 'Златоуст ТРК Тарелка', 'Златоуст',
                 'Артиллерийская', 'Гагарина', 'Копейск', 'КС Теплотех', 'Сталеваров', 'Худякова', 'Комсомольский',
                 'Молодогвардейцев', 'Ленина']
    df = df.reindex(columns=['Номенклатура', 'Продажи', "Маркса", 'Склад'] + prioritet)

    conditions = [
        df['Продажи'] - df['Маркса'] > 0,
        df['Продажи'] - df['Маркса'] < 0,
        df['Продажи'] - df['Маркса'] == 0
    ]

    choices = [
        df['Продажи'] - df['Маркса'],  # если положительное
        0,  # если отрицательное
        1  # если ноль
    ]

    df.insert(2, 'Рекомендовано к заказу', np.select(conditions, choices, 0))
    df['ordered'] = False
    df = df[(df['Рекомендовано к заказу'] > 0)]

    for idx in df.index:
        if df.loc[idx, "Склад"] >= df.loc[idx, "Рекомендовано к заказу"]:
            df.loc[idx, 'ordered'] = True
        elif 0 < df.loc[idx, "Склад"] < df.loc[idx, "Рекомендовано к заказу"]:
            df.loc[idx, "Рекомендовано к заказу"] = df.loc[idx, "Склад"]
            df.loc[idx, 'ordered'] = True

    result = df[(df['ordered'] == True)]
    output_file = output_dir / f'Склад Запчасти от {current_date}.xlsx'
    result[['Номенклатура', "Рекомендовано к заказу", 'Маркса', 'Склад']].to_excel(output_file, index=False)
    redactor(output_file)
    print(f"создан файл 'заказы со склада (запчасти) от {current_date}.xlsx' найдено {len(result)} позиций")

    df = df[(df['ordered'] == False)]
    df.insert(2, 'Заказ из магазина', 0)
    df['ordered'] = 0

    for idx in df.index:
        for i in range(len(prioritet)):
            warehouse = prioritet[i]
            if df.loc[idx, 'Маркса'] > 0:
                if df.loc[idx, warehouse] > 2 and df.loc[idx, "Рекомендовано к заказу"] > df.loc[idx, "ordered"] and \
                        df.loc[idx, warehouse] - df.loc[idx, 'Маркса'] >= 2:
                    df.loc[idx, "ordered"] += 1
                    df.loc[idx, "Заказ из магазина"] = 1
                else:
                    df.loc[idx, warehouse] = 0
            else:
                if warehouse == 'Ленина':
                    if df.loc[idx, warehouse] > 2 and df.loc[idx, "Рекомендовано к заказу"] > df.loc[
                        idx, "ordered"]:  # Проверяем наличие товара
                        # Устанавливаем 1 на найденном складе
                        df.loc[idx, "ordered"] += 1
                        df.loc[idx, "Заказ из магазина"] = 1
                    else:
                        df.loc[idx, warehouse] = 0
                else:
                    if df.loc[idx, warehouse] > 1 and df.loc[idx, "Рекомендовано к заказу"] > df.loc[idx, "ordered"]:
                        df.loc[idx, "ordered"] += 1
                        df.loc[idx, "Заказ из магазина"] = 1
                    else:
                        df.loc[idx, warehouse] = 0



    # Создаем новый workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Создаем листы для каждого склада
    for sklad in prioritet:
        # Фильтруем строки, где этот склад имеет значение 1
        otchet = df[(df[sklad] > 1)]

        # Проверяем, есть ли данные для этого склада
        if not otchet.empty:
            # Создаем новый лист с названием склада
            ws = wb.create_sheet(title=sklad)

            # Добавляем данные
            for r in dataframe_to_rows(
                    otchet[['Номенклатура', "Заказ из магазина", "Рекомендовано к заказу", "Маркса", sklad, "ordered"]],
                    index=False, header=True):
                ws.append(r)

            redactor_ws(ws)

        else:
            print(f"Нет данных для склада {sklad}")

    # Сохраняем файл
    filename = output_dir / f'Магазины Запчасти от {current_date}.xlsx'
    wb.save(filename)
    print(f"Создан файл '{filename}' с {len(wb.sheetnames)} листами")

    # df = df[(df['ordered'] < df["Рекомендовано к заказу"])]
    # df = df[(df['Продажи'] > 0)]
    # df[['Номенклатура', 'Продажи', 'ordered']].to_excel(f'дефицит от {current_date}.xlsx', index=False)
    # redactor(f'дефицит от {current_date}.xlsx')
    # print(f'Cоздан файл - дефицит запчастей от {current_date}.xlsx')

    end_time = time.perf_counter()
    execution_time = end_time - start_time
    print(f'Отчет завершен, обработано {dlina} позиций за {execution_time:.4f} секунд')

if __name__ == "__main__":
    asyncio.run(zip_otchet())
