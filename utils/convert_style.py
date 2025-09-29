from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Загрузка существующего файла или создание нового
def redactor(file):
    try:
        wb = load_workbook(file)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active

    # 1. Задание размера колонок
    # Автоподбор ширины
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column_letter].width = adjusted_width

    # Или фиксированная ширина
    # ws.column_dimensions['A'].width = 20
    # ws.column_dimensions['B'].width = 30

    # 2. Чередование цветов строк
    # Создаем заполнения
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Применяем к строкам (начиная со второй строки, если первая - заголовки)
    for row in range(2, ws.max_row + 1):  # Начинаем с 2, чтобы пропустить заголовки
        if row % 2 == 0:  # Четные строки
            fill = light_gray_fill
        else:  # Нечетные строки
            fill = white_fill

        # Применяем к каждой ячейке в строке
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill

    # 3. Форматирование заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:  # Первая строка - заголовки
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 4. Сохранение файла
    wb.save(file)
    print("Файл успешно отредактирован!")

def redactor_ws(ws):
    """
    Редактирует лист Excel, применяя форматирование
    ws - объект worksheet openpyxl
    """
    # Задание размера колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Форматирование заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for cell in ws[1]:  # Первая строка - заголовки
        cell.font = header_font
        cell.fill = header_fill

    # Чередование цветов строк
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill